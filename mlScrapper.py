import requests
from bs4 import BeautifulSoup
from lxml import etree
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os
#from PIL import Image

print('Escribe lo que deseas buscar')
searchTerm = input('> ')
print(f'Buscando: {searchTerm}')
searchTerm = searchTerm.replace(' ','-')

base_url = 'https://listado.mercadolibre.com.mx/'
new_items_filter = "_ITEM*CONDITION_2230284_NoIndex_True#applied_filter_id%3DITEM_CONDITION%26applied_filter_name%3DCondición%26applied_filter_order%3D9%26applied_value_id%3D2230284%26applied_value_name%3DNuevo%26applied_value_order%3D1%26applied_value_results%3D78%26is_custom%3Dfalse"

r = requests.get(f'{base_url}+{searchTerm}+{new_items_filter}')
#Status code must be 200
print('status:', r.status_code)
soup = BeautifulSoup(r.content, 'html.parser')

#Start getting atributes in the search page

titles = soup.find_all('h2', attrs={'class': 'ui-search-item__title'})
titles = [i.text for i in titles]
print('titles:', len(titles))

""" urls = soup.find_all('a', attrs={'class':'ui-search-result__image'})
urls = [i.get('href') for i in urls]
print('url´s:',  len(urls)) """

#create a dom element with etree to use the Xpath and find difficult elements
dom  = etree.HTML(str(soup))

urls = dom.xpath("//li[@class='ui-search-layout__item']//div[@class='ui-search-result__image']//a[@class='ui-search-link']")
print('urls:',  len(urls))
urls = [i.attrib['href'] for i in urls]

prices = dom.xpath("//li[@class='ui-search-layout__item']//div[@class='ui-search-result__content-wrapper']//div[@class='ui-search-item__group__element ui-search-price__part-without-link']//div[@class='ui-search-price__second-line']//span[@class='price-tag-fraction']")
print('prices:',  len(prices))
prices = [int(i.text.replace(',','')) for i in prices]

images_url = dom.xpath("//li[@class='ui-search-layout__item']//div[@class='ui-search-result__image']//div[@class='slick-track']//img[1]")
images_url = [i.attrib['data-src'] for i in images_url]
print('image url´s:', len(images_url))

# use this two prints to check if you are getting the correct urls
# print(dir(images_url[0]))
#print(images_url[0].attrib)
#print(images_url[0].attrib['data-src'])

#Create a dataset and extract it to csv
df = pd.DataFrame({"titulos":titles, "img": images_url, "url":urls, "precios": prices} )
df = df.set_index('titulos')

with pd.ExcelWriter('DataFrame.xlsx') as writer:
    df.to_excel(writer, sheet_name='database')

wb = load_workbook('DataFrame.xlsx')
ws = wb['database']

def download_img(url,file_name):
    res = requests.get(url)
    img = open(file_name, 'wb')
    img.write(res.content)
    img.close()

def insert_img(file_name,cell):
    img = Image(file_name)
    img.width,img.height=72,72
    ws.add_image(img, cell)

def remove_img(img_name):
# check if file exists or not
    if os.path.exists(img_name) is False:
        # file did not exists
        return True
    else:
        os.remove( img_name)
        

for i in range(2,len(images_url)-2):
    name = str(i)#ws['B'+str(i)].value.split('/')[3].split('.')[0]
    url = ws['B'+str(i)].value
    print('extrayendo', i-1, " de ", len(images_url)-2)
    download_img(url,name)
    ws['B'+str(i)]=""
    ws.row_dimensions[i].height=80
    ws.row_dimensions[i].width=80
    insert_img(name,'B'+str(i))
    remove_img(name)

""" worksheet = wb.getWorksheets().get(0)
cells = worksheet.getCells()
cells.setColumnWidth(1, 40)
cells.setColumnWidth(2, 80)
cells.setColumnWidth(1, 80)
cells.setColumnWidth(1, 80) """
#os.remove(file) for filename in os.listdir('/') if file.startswith('D_NQ_NP')
wb.save('output.xlsx') 