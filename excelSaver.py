from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import alignment
import requests
import os

def imageAdder(len):
    print('Descargando imágenes ....')
    #wb = load_workbook('DeepScrapping.xlsx')
    wb = load_workbook('DataFrame.xlsx')
    ws = wb['database']

    def download_img(url,file_name):
        res = requests.get(url)
        img = open(file_name, 'wb')
        img.write(res.content)
        img.close()

    def insert_img(file_name,cell):
        img = Image(file_name)
        img.width,img.height=100,100
        ws.add_image(img, cell)

    def remove_img(img_name):
        try:
            os.remove(img_name)
        except:
            pass
    # check if file exists or not
        #if os.path.exists(img_name) is False:
            # file did not exists
            #return True

    for i in range(2,len):
        name = ws['B'+str(i)].value.split('/')[3].split('.')[0]
        print('Descargando imagen ' + str(i-1) + ' de ' + str(len-2))
        url = ws['B'+str(i)].value
        download_img(url,name)
        ws['B'+str(i)]=""
        ws.row_dimensions[i].height=80
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 5
        insert_img(name,'B'+str(i))
    wb.save('output.xlsx') 
    wb = load_workbook('DataFrame.xlsx')
    ws = wb['database']

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment =  cell.alignment.copy(wrapText=True, horizontal='general',vertical='bottom',shrink_to_fit=False)
    for i in range(2,len):
        name = ws['B'+str(i)].value.split('/')[3].split('.')[0]
        remove_img(name)
    print('Eliminando imágenes descargadas')
    print('Archivo final guardado')
